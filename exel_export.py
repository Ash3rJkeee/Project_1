# coding: utf8

import openpyxl
import YaParser
import gismeteo_parser
import Meteoinfo_parser
import WeatherCom_parser
import datetime
import smart_request
from openpyxl.styles import Alignment

"""Модуль изымает данные из модулей парсеров и сводит их в общую exel таблицу"""

# todo Включить обработку

def date_to_exel_format(day):
    return day.strftime('%d.%m.%Y')

def raise_exel_date(string):
    """Увеличивает дату в формате Exel на 1 день"""
    a = datetime.datetime.strptime(string, "%d.%m.%Y") + datetime.timedelta(days=1)
    return a.strftime("%d.%m.%Y")

def export():
    try:
        if smart_request.checked == True:
            data1 = [YaParser.temps_night[0], YaParser.temps_day[0],
                    gismeteo_parser.temps_night[0], gismeteo_parser.temps_day[0],
                    Meteoinfo_parser.temps_night[0], Meteoinfo_parser.temps_day[0],
                    WeatherCom_parser.temps_night[0], WeatherCom_parser.temps_day[0]]

            data2 = [YaParser.temps_night[1], YaParser.temps_day[1],
                    gismeteo_parser.temps_night[1], gismeteo_parser.temps_day[1],
                    Meteoinfo_parser.temps_night[1], Meteoinfo_parser.temps_day[1],
                    WeatherCom_parser.temps_night[1], WeatherCom_parser.temps_day[1]]

            data3 = [YaParser.temps_night[2], YaParser.temps_day[2],
                    gismeteo_parser.temps_night[2], gismeteo_parser.temps_day[2],
                    Meteoinfo_parser.temps_night[2], Meteoinfo_parser.temps_day[2],
                    WeatherCom_parser.temps_night[2], WeatherCom_parser.temps_day[2]]
        else:
            print('Соединение не установлено. Экспорт невозможен.')
            raise Exception("Соединение не установлено. Экспорт невозможен.")
    except IndexError:
        print('Не сформирован массив загрузки. Запустите все парсеры.')
        raise FileExistsError('Не сформирован массив загрузки. Запустите все парсеры.')

    # data1 = ['-15', '24', '16', 22, '14.7', '25.0', '15', '25']
    # data2 = ['16', '22', '13', '28', '-12.7', '26.0', '14', '26']
    # data3 = ['11', '23', '13', '26', '15.7', '27.0', '14', '30']

    print(data1)
    print(data2)
    print(data3)

    print('*******************************************************')
    print('Идет преобразование и экспорт в файл')

    # преобразование не правильных "-" а так же преобразование к числам для excel
    for i in range(len(data1)):
        if data1[i] != "ошибка":
            if not str(data1[i])[0].isdigit():                           # если 1 символ знак тире и минус
                data1[i] = float(str(data1[i])[1:]) * (-1)                 # сменить знак
            else:
                data1[i] = float(data1[i])
    print(data1)

    for i in range(len(data2)):
        if data2[i] != "ошибка":
            if not str(data2[i])[0].isdigit():
                data2[i] = float(str(data2[i])[1:]) * (-1)
            else:
                data2[i] = float(data2[i])
    print(data2)

    for i in range(len(data3)):
        if data3[i] != "ошибка":
            if not str(data3[i])[0].isdigit():
                data3[i] = float(str(data3[i])[1:]) * (-1)
            else:
                data3[i] = float(data3[i])
    print(data3)



    # file = r'Прогнозы.xlsx'
    file = r'\\tsk-fileserv-01.tsk-mosenergo.ru\ОИО\ТСК-Мосэнерго (АУ)\Центральная диспетчерская служба\Область обмена\ЦДС\Аналитика\Прогнозы.xlsx'

    wb = openpyxl.load_workbook(file)
    # wb_lists = wb.sheetnames
    # print(wb_lists)
    ws = wb[wb.sheetnames[0]]
    # print(type(ws.max_row))

    # сегодняшняя дата (для отладки пробую имитировать завтрашнюю дату, путем смещения текущей)
    today = datetime.date.today() + datetime.timedelta(days=0)
    # print(today)

    # делаем отметку с новой датой записи прогноза
    # получаем список всех дат
    # date_list_2 = list(ws['B'])
    # date_list_2 = date_list_2[3:]

    date_list_1 = list(ws['A'])
    date_list_1 = date_list_1[3:]

    for i in range(len(date_list_1)):
        date_list_1[i] = date_list_1[i].value
        # print(date_list_1[i])

    # готовим записи для дат прогнозов
    mark_row = 3
    while ws.cell(row=mark_row, column=2).value != date_to_exel_format(today + datetime.timedelta(days=3)):
        mark_row += 1
        if ws.cell(row=mark_row, column=2).value is None:
            ws.cell(row=mark_row, column=2).value = raise_exel_date(ws.cell(row=mark_row-1, column=2).value)
            ws.cell(row=mark_row, column=2).alignment = Alignment(horizontal="center", vertical="center")     # выравнивание

    mark_row = mark_row - 2



    # задаем номера строчек для последующей записи
    if date_to_exel_format(today) not in date_list_1:
        print('Сделана новая запись на', today)
        ws.cell(row=mark_row, column=1).value = date_to_exel_format(today)              # если запись от текущей даты не делалась
        ws.cell(row=mark_row, column=1).alignment = Alignment(horizontal="center", vertical="center")    # выравнивание
    else:
        print('Данные на текущую дату перезаписаны')
        mark_row = date_list_1.index(date_to_exel_format(today)) + 4

    forecast_1_day_row = mark_row
    forecast_2_days_row = mark_row + 1
    forecast_3_days_row = mark_row + 2

    i = 0
    for col in range(3, 11):
        ws.cell(row=forecast_1_day_row, column=col).value = data1[i]
        ws.cell(row=forecast_1_day_row, column=col).alignment = Alignment(horizontal="center", vertical="center") # выравнивание
        i = i + 1

    i = 0
    for col in range(11, 19):
        ws.cell(row=forecast_2_days_row, column=col).value = data2[i]
        ws.cell(row=forecast_2_days_row, column=col).alignment = Alignment(horizontal="center", vertical="center") # выравнивание
        i = i + 1

    i = 0
    for col in range(19, 27):
        ws.cell(row=forecast_3_days_row, column=col).value = data3[i]
        ws.cell(row=forecast_3_days_row, column=col).alignment = Alignment(horizontal="center", vertical="center") # выравнивание
        i = i + 1

    wb.save(file)


if __name__ == '__main__':
    try:
        YaParser.yaParser()
    except:
        for i in [0, 1, 2]:
            YaParser.temps_night[i] = "ошибка"
            YaParser.temps_day[i] = "ошибка"
    try:
        gismeteo_parser.gismeteo_parser()
    except:
        for i in [0, 1, 2]:
            gismeteo_parser.temps_night[i] = "ошибка"
            gismeteo_parser.temps_day[i] = "ошибка"
    try:
        Meteoinfo_parser.parser()
    except:
        for i in [0, 1, 2]:
            Meteoinfo_parser.temps_night[i] = "ошибка"
            Meteoinfo_parser.temps_day[i] = "ошибка"
    try:
        WeatherCom_parser.parser()
    except:
        for i in [0, 1, 2]:
            WeatherCom_parser.temps_night[i] = "ошибка"
            WeatherCom_parser.temps_day[i] = "ошибка"
    export()

    # todo сделать задержку закрытия консоли в течение 5 сек если не будет нажатия клавиатуры
    # input('Нажмите Enter для выхода')
