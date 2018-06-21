"""Модуль для организации программы с exel файлом"""

import openpyxl, datetime

# todo Добавить проверку наличия записи на дату. Чтобы исключить многократный прогноз.
# todo написать функцию преобразования даты из "гггг-мм-дд" в "дд.мм.гггг"


def date_to_exel_format(day):
    return day.strftime('%d.%m.%Y')


file = 'Forecasts.xlsx'
wb = openpyxl.load_workbook(file)

wb_lists = wb.sheetnames
print(wb_lists)

ws = wb['Лист1']

ws['A4'] = date_to_exel_format(datetime.date.today())



wb.save(file)