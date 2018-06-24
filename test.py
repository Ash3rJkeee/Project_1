import openpyxl
import YaParser
import gismeteo_parser
import Meteoinfo_parser
import WeatherCom_parser
import datetime
"""Модуль изымает данные из модулей парсеров и сводит их в общую exel таблицу"""


def date_to_exel_format(day):
    return day.strftime('%d.%m.%Y')


YaParser.yaParser()
gismeteo_parser.gismeteo_parser()
Meteoinfo_parser.parser()
WeatherCom_parser.parser()


data1 = [YaParser.temps_night[0], YaParser.temps_day[0],
        gismeteo_parser.temps_night[0], gismeteo_parser.temps_day[0],
        Meteoinfo_parser.temps_night[0], Meteoinfo_parser.temps_day[0],
        WeatherCom_parser.temps_night[0], WeatherCom_parser.temps_day[0]]

data2 = [YaParser.temps_night[1], YaParser.temps_day[1],
        gismeteo_parser.temps_night[1], gismeteo_parser.temps_day[1],
        Meteoinfo_parser.temps_night[1], Meteoinfo_parser.temps_day[1],
        WeatherCom_parser.temps_night[1], WeatherCom_parser.temps_day[1]]

data3 = [YaParser.temps_night[2], YaParser.temps_day[2],
        gismeteo_parser.temps_night[2], gismeteo_parser.temps_day[2],
        Meteoinfo_parser.temps_night[2], Meteoinfo_parser.temps_day[2],
        WeatherCom_parser.temps_night[2], WeatherCom_parser.temps_day[2]]


# data1 = ['15', '24', '16', '22', '14.7', '25.0', '15', '25']
# data2 = ['16', '22', '13', '28', '12.7', '26.0', '14', '26']
# data3 = ['11', '23', '13', '26', '15.7', '27.0', '14', '30']

print('*******************************************************')
print(data1)
print(data2)
print(data3)

file = 'Forecasts.xlsx'
wb = openpyxl.load_workbook(file)
wb_lists = wb.sheetnames
# print(wb_lists)
ws = wb[wb.sheetnames[0]]
# print(type(ws.max_row))

# сегодняшняя дата (пробую имитировать завтрашнюю дату, путем смещения текущей)
today = datetime.date.today() + datetime.timedelta(days=0)
print(today)

# делаем отметку с новой датой записи прогноза
# получаем список всех дат
date_list = list(ws['A'])
date_list = date_list[3:]

empty_A_row = 0
for i in range(len(date_list)):
    if date_list[i].value is not None:
        # если в таблицу потом будет писаться время, то тут будет date_list[i] = date_list[i].value.date()
        # для комфортного преобразования из типа datetime.datetime
        date_list[i] = date_list[i].value
    else:
        date_list[i] = date_list[i].value
        empty_A_row = i + 4                      # сразу найдем, номер следующей пустой строки для даты записи
        # print('empty_A_row', empty_A_row)
        break
    # print(date_list[i])

# задаем номера строчек для последующей записи
if date_to_exel_format(today) not in date_list:
    print('today not in date_list')
    today_row = empty_A_row
    ws.cell(row=today_row, column=1).value = date_to_exel_format(today)              # если запись не делалась, сделать заготовку
else:
    print('today in date_list')
    today_row = date_list.index(date_to_exel_format(today)) + 4

forecast_1_day_row = today_row
forecast_2_days_row = today_row + 1
forecast_3_days_row = today_row + 2

# готовим запись для дат прогнозов
for i in [0, 1, 2]:
    ws.cell(row=today_row + i, column=2).value = date_to_exel_format(today + datetime.timedelta(days=i + 1))

i = 0
for col in range(3, 11):
    ws.cell(row=forecast_1_day_row, column=col).value = float(data1[i])
    i = i + 1

i = 0
for col in range(11, 19):
    ws.cell(row=forecast_2_days_row, column=col).value = float(data2[i])
    i = i + 1

i = 0
for col in range(19, 27):
    ws.cell(row=forecast_3_days_row, column=col).value = float(data3[i])
    i = i + 1

wb.save(file)
